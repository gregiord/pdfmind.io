"""
Paperless conversion backend.

One small Flask service that powers the six tools that can't run in the browser:
  - office-to-pdf : Word / Excel / PowerPoint  ->  PDF   (LibreOffice headless)
  - pdf-to-word   : PDF  ->  .docx                       (pdf2docx)
  - pdf-to-excel  : PDF tables  ->  .xlsx                (pdfplumber + openpyxl)
  - protect       : PDF + password  ->  encrypted PDF    (pikepdf)

Everything else in the toolkit runs client-side and never touches this server.
Run it with Docker (see README) or locally with:  python app.py
"""

import io
import os
import shutil
import subprocess
import tempfile
import uuid

from flask import Flask, request, send_file, jsonify
from flask_cors import CORS

app = Flask(__name__)

# Allow the static front-end (served from a different origin) to call us.
# For production, replace "*" with your site's URL, e.g. "https://paperless.example.com".
CORS(app, resources={r"/api/*": {"origins": os.environ.get("ALLOW_ORIGIN", "*")}})

SOFFICE = os.environ.get("SOFFICE_BIN", "soffice")
MAX_MB = int(os.environ.get("MAX_UPLOAD_MB", "50"))
app.config["MAX_CONTENT_LENGTH"] = MAX_MB * 1024 * 1024


def _err(message, code=400):
    return jsonify(error=message), code


def _save_upload(workdir, default_name):
    f = request.files.get("file")
    if not f or f.filename == "":
        return None, _err("No file was uploaded.")
    safe = os.path.basename(f.filename) or default_name
    path = os.path.join(workdir, safe)
    f.save(path)
    return path, None


def _send_bytes(path, download_name, mimetype):
    """Read the result into memory so we can delete the temp dir before returning."""
    with open(path, "rb") as fh:
        data = fh.read()
    return send_file(
        io.BytesIO(data),
        as_attachment=True,
        download_name=download_name,
        mimetype=mimetype,
    )


@app.get("/")
def health():
    return jsonify(status="ok", service="paperless-backend")


@app.post("/api/office-to-pdf")
def office_to_pdf():
    """Word / Excel / PowerPoint -> PDF via LibreOffice headless."""
    workdir = tempfile.mkdtemp(prefix="lo_")
    try:
        src, err = _save_upload(workdir, "input")
        if err:
            return err
        # A unique user profile avoids the "soffice is already running" lock
        # when several conversions happen at once.
        profile = "file://" + os.path.join(workdir, "profile")
        proc = subprocess.run(
            [
                SOFFICE,
                "-env:UserInstallation=" + profile,
                "--headless",
                "--norestore",
                "--convert-to",
                "pdf",
                "--outdir",
                workdir,
                src,
            ],
            capture_output=True,
            timeout=180,
        )
        out = os.path.splitext(src)[0] + ".pdf"
        if not os.path.exists(out):
            return _err("LibreOffice could not convert this file. " +
                        proc.stderr.decode("utf-8", "ignore")[:300], 500)
        return _send_bytes(out, "converted.pdf", "application/pdf")
    except subprocess.TimeoutExpired:
        return _err("Conversion timed out.", 504)
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


@app.post("/api/pdf-to-word")
def pdf_to_word():
    """PDF -> editable .docx via pdf2docx."""
    from pdf2docx import Converter

    workdir = tempfile.mkdtemp(prefix="p2w_")
    try:
        src, err = _save_upload(workdir, "input.pdf")
        if err:
            return err
        out = os.path.join(workdir, "out.docx")
        cv = Converter(src)
        try:
            cv.convert(out)  # all pages
        finally:
            cv.close()
        if not os.path.exists(out):
            return _err("Could not convert this PDF to Word.", 500)
        return _send_bytes(
            out,
            "converted.docx",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


@app.post("/api/pdf-to-excel")
def pdf_to_excel():
    """PDF tables -> .xlsx (best effort) via pdfplumber + openpyxl."""
    import pdfplumber
    from openpyxl import Workbook

    workdir = tempfile.mkdtemp(prefix="p2x_")
    try:
        src, err = _save_upload(workdir, "input.pdf")
        if err:
            return err
        out = os.path.join(workdir, "out.xlsx")
        wb = Workbook()
        wb.remove(wb.active)
        found = False
        with pdfplumber.open(src) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                ws = wb.create_sheet(title=f"Page {i}"[:31])
                row = 1
                for table in page.extract_tables() or []:
                    found = True
                    for r in table:
                        for c, val in enumerate(r, start=1):
                            ws.cell(row=row, column=c, value=val)
                        row += 1
                    row += 1  # blank line between tables
        if not found:
            ws = wb.create_sheet(title="Sheet1")
            ws["A1"] = "No tables were detected in this PDF."
        wb.save(out)
        return _send_bytes(
            out,
            "converted.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


@app.post("/api/protect")
def protect():
    """Encrypt a PDF with a password via pikepdf (AES-256)."""
    import pikepdf

    password = request.form.get("password", "")
    if not password:
        return _err("A password is required.")
    workdir = tempfile.mkdtemp(prefix="prot_")
    try:
        src, err = _save_upload(workdir, "input.pdf")
        if err:
            return err
        out = os.path.join(workdir, "protected.pdf")
        with pikepdf.open(src) as pdf:
            pdf.save(
                out,
                encryption=pikepdf.Encryption(owner=password, user=password, R=6),
            )
        return _send_bytes(out, "protected.pdf", "application/pdf")
    except pikepdf._core.PasswordError:
        return _err("That PDF is already password-protected; unlock it first.", 400)
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8080"))
    app.run(host="0.0.0.0", port=port)
